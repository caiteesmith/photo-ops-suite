# =========================================
# file: tools/finance_dashboard.py
# =========================================
from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple
from io import BytesIO

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

# -------------------------
# Config
# -------------------------
@dataclass(frozen=True)
class FinanceSheetConfig:
    sheet_name: str

    # Headline cells (Monthly Cash Flow row)
    net_income_cell: str
    mortgage_cell: str
    expenses_cell: str
    roth_savings_cell: str
    remaining_cell: str

    # Net Worth total cell
    net_worth_total_cell: str

    # Tables (top-left anchored ranges)
    # Each range should include headers in the first row.
    income_table_range: str
    savings_table_range: str
    expenses_table_range: str
    investments_table_range: str
    net_worth_table_range: str


DEFAULT_CONFIG = FinanceSheetConfig(
    sheet_name="Finances",
    # ⚠️ These are BEST-GUESS defaults based on your screenshot.
    # If something comes back blank, just adjust the cell refs.
    net_income_cell="C4",
    mortgage_cell="E4",
    expenses_cell="G4",
    roth_savings_cell="I4",
    remaining_cell="K4",
    net_worth_total_cell="K21",
    # Ranges are also best-guess; tweak to match your exact workbook.
    income_table_range="A7:C15",
    savings_table_range="A17:C23",
    expenses_table_range="F7:H24",
    investments_table_range="I7:K12",
    net_worth_table_range="I16:K21",
)


# -------------------------
# Excel helpers
# -------------------------
def _cell_value(ws: Worksheet, addr: str) -> Any:
    return ws[addr].value


def _read_range_as_df(ws: Worksheet, excel_range: str) -> pd.DataFrame:
    """
    Reads a rectangular Excel range into a DataFrame.
    First row is treated as headers.
    """
    cells = ws[excel_range]
    rows = [[c.value for c in row] for row in cells]
    if not rows or len(rows) < 2:
        return pd.DataFrame()

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    data = rows[1:]

    df = pd.DataFrame(data, columns=headers)

    # Drop empty columns and fully-empty rows
    df = df.dropna(axis=1, how="all")
    df = df.dropna(axis=0, how="all")

    # Normalize column names
    df.columns = [str(c).strip() for c in df.columns]
    return df


def _to_float(x: Any) -> Optional[float]:
    if x is None:
        return None
    try:
        return float(x)
    except Exception:
        return None


def _money(x: Optional[float]) -> str:
    if x is None:
        return "—"
    return f"${x:,.2f}"


# -------------------------
# Core loader
# -------------------------
from io import BytesIO

def load_finance_workbook(file_bytes: bytes, cfg: FinanceSheetConfig) -> Dict[str, Any]:
    wb = load_workbook(filename=BytesIO(file_bytes), data_only=True)
    if cfg.sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{cfg.sheet_name}' not found. Available: {wb.sheetnames}")

    ws = wb[cfg.sheet_name]

    headline = {
        "Net Income": _to_float(_cell_value(ws, cfg.net_income_cell)),
        "Mortgage": _to_float(_cell_value(ws, cfg.mortgage_cell)),
        "Expenses": _to_float(_cell_value(ws, cfg.expenses_cell)),
        "Roth/Savings": _to_float(_cell_value(ws, cfg.roth_savings_cell)),
        "Remaining": _to_float(_cell_value(ws, cfg.remaining_cell)),
    }

    net_worth_total = _to_float(_cell_value(ws, cfg.net_worth_total_cell))

    tables = {
        "Income": _read_range_as_df(ws, cfg.income_table_range),
        "Savings": _read_range_as_df(ws, cfg.savings_table_range),
        "Expenses": _read_range_as_df(ws, cfg.expenses_table_range),
        "Investments": _read_range_as_df(ws, cfg.investments_table_range),
        "Net Worth": _read_range_as_df(ws, cfg.net_worth_table_range),
    }

    return {
        "headline": headline,
        "net_worth_total": net_worth_total,
        "tables": tables,
        "loaded_at": datetime.now().isoformat(timespec="seconds"),
        "sheet": cfg.sheet_name,
        "config": cfg,
    }


# -------------------------
# UI
# -------------------------
def render_personal_finance_dashboard():
    st.title("💸 Personal Finance Dashboard")
    st.caption(
        "Upload your finance spreadsheet to get a clean dashboard view (cash flow, expenses, investments, net worth). "
        "This reads your sheet’s *existing calculated values*."
    )

    with st.expander("How this works (important)", expanded=False):
        st.markdown(
            """
- This tool **does not recalculate Excel formulas** the way Excel does.
- It reads the **last saved values** from the workbook.
- If you see blanks/zeros, open the spreadsheet in Excel/Numbers, let it calculate, **save**, then upload again.
            """.strip()
        )

    uploaded = st.file_uploader("Upload your spreadsheet (.xlsx)", type=["xlsx"], key="finance_upload")
    if not uploaded:
        st.info("Upload your .xlsx file to begin.")
        return

    # Config editor (so you can dial in your exact sheet/cells without touching code)
    with st.expander("Config (adjust if your layout changes)", expanded=False):
        cfg = FinanceSheetConfig(
            sheet_name=st.text_input("Sheet name", value=DEFAULT_CONFIG.sheet_name, key="cfg_sheet"),
            net_income_cell=st.text_input("Net Income cell", value=DEFAULT_CONFIG.net_income_cell, key="cfg_net_income"),
            mortgage_cell=st.text_input("Mortgage cell", value=DEFAULT_CONFIG.mortgage_cell, key="cfg_mortgage"),
            expenses_cell=st.text_input("Expenses cell", value=DEFAULT_CONFIG.expenses_cell, key="cfg_expenses"),
            roth_savings_cell=st.text_input("Roth/Savings cell", value=DEFAULT_CONFIG.roth_savings_cell, key="cfg_roth"),
            remaining_cell=st.text_input("Remaining cell", value=DEFAULT_CONFIG.remaining_cell, key="cfg_remaining"),
            net_worth_total_cell=st.text_input("Net Worth TOTAL cell", value=DEFAULT_CONFIG.net_worth_total_cell, key="cfg_nw_total"),
            income_table_range=st.text_input("Income table range", value=DEFAULT_CONFIG.income_table_range, key="cfg_income_tbl"),
            savings_table_range=st.text_input("Savings table range", value=DEFAULT_CONFIG.savings_table_range, key="cfg_savings_tbl"),
            expenses_table_range=st.text_input("Expenses table range", value=DEFAULT_CONFIG.expenses_table_range, key="cfg_expenses_tbl"),
            investments_table_range=st.text_input("Investments table range", value=DEFAULT_CONFIG.investments_table_range, key="cfg_invest_tbl"),
            net_worth_table_range=st.text_input("Net Worth table range", value=DEFAULT_CONFIG.net_worth_table_range, key="cfg_nw_tbl"),
        )

    try:
        data = load_finance_workbook(uploaded.getvalue(), cfg)
    except Exception as e:
        st.error(f"Could not load workbook: {e}")
        return

    headline = data["headline"]
    net_worth_total = data["net_worth_total"]
    tables: Dict[str, pd.DataFrame] = data["tables"]

    # ----- Headline metrics -----
    st.subheader("Monthly Cash Flow")
    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("Net Income", _money(headline.get("Net Income")))
    c2.metric("Mortgage", _money(headline.get("Mortgage")))
    c3.metric("Expenses", _money(headline.get("Expenses")))
    c4.metric("Roth/Savings", _money(headline.get("Roth/Savings")))
    c5.metric("Remaining", _money(headline.get("Remaining")))

    st.subheader("Net Worth")
    st.metric("Total Net Worth", _money(net_worth_total))

    st.divider()

    # ----- Tabs for tables -----
    tab1, tab2, tab3, tab4, tab5 = st.tabs(["Expenses", "Income", "Savings", "Investments", "Net Worth"])

    with tab1:
        df = tables.get("Expenses", pd.DataFrame())
        if df.empty:
            st.warning("Expenses table is empty (check your range in Config).")
        else:
            st.dataframe(df, use_container_width=True, hide_index=True)

            # Try to chart if a numeric column exists
            numeric_cols = [c for c in df.columns if df[c].map(lambda v: isinstance(v, (int, float))).any()]
            if numeric_cols:
                amount_col = numeric_cols[-1]
                label_col = df.columns[0]
                chart_df = df[[label_col, amount_col]].dropna()
                chart_df = chart_df[chart_df[amount_col].map(lambda v: isinstance(v, (int, float)))]
                chart_df = chart_df.sort_values(by=amount_col, ascending=False).head(20)
                st.bar_chart(chart_df.set_index(label_col)[amount_col], use_container_width=True)

    with tab2:
        df = tables.get("Income", pd.DataFrame())
        if df.empty:
            st.warning("Income table is empty (check your range in Config).")
        else:
            st.dataframe(df, use_container_width=True, hide_index=True)

    with tab3:
        df = tables.get("Savings", pd.DataFrame())
        if df.empty:
            st.warning("Savings table is empty (check your range in Config).")
        else:
            st.dataframe(df, use_container_width=True, hide_index=True)

    with tab4:
        df = tables.get("Investments", pd.DataFrame())
        if df.empty:
            st.warning("Investments table is empty (check your range in Config).")
        else:
            st.dataframe(df, use_container_width=True, hide_index=True)

    with tab5:
        df = tables.get("Net Worth", pd.DataFrame())
        if df.empty:
            st.warning("Net Worth table is empty (check your range in Config).")
        else:
            st.dataframe(df, use_container_width=True, hide_index=True)

    st.divider()

    # ----- Downloads -----
    st.subheader("Export")
    export_payload = {
        "loaded_at": data["loaded_at"],
        "sheet": data["sheet"],
        "headline": headline,
        "net_worth_total": net_worth_total,
        "tables": {k: v.to_dict(orient="records") for k, v in tables.items()},
        "config": cfg.__dict__,
    }

    colA, colB = st.columns(2)
    with colA:
        st.download_button(
            "Download JSON snapshot",
            data=pd.Series(export_payload).to_json(indent=2),
            file_name="finance_dashboard_snapshot.json",
            mime="application/json",
            use_container_width=True,
        )
    with colB:
        # One combined CSV of the big tables (Expenses + Investments) is often most useful
        combined = []
        for name in ["Expenses", "Investments", "Net Worth"]:
            df = tables.get(name)
            if df is not None and not df.empty:
                tmp = df.copy()
                tmp.insert(0, "Table", name)
                combined.append(tmp)
        if combined:
            out_df = pd.concat(combined, ignore_index=True)
            st.download_button(
                "Download tables (CSV)",
                data=out_df.to_csv(index=False),
                file_name="finance_tables.csv",
                mime="text/csv",
                use_container_width=True,
            )
        else:
            st.info("No tables available to export yet (check ranges).")